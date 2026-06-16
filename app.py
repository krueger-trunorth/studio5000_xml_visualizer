"""Plotly Dash app to browse the exploded/ XML tree.

Run with:  python app.py
"""

import base64
import io
import re
import xml.dom.minidom as minidom
import xml.etree.ElementTree as ET
from pathlib import Path

import dash
import dash_mantine_components as dmc
import diskcache
import pandas as pd
from dash import DiskcacheManager, Input, Output, State, dcc, html
from dash_iconify import DashIconify
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from acd_pipeline import PipelineError, import_acd
from xml_to_csv import (
    FAULT_FIELDS,
    fault_row,
    recipe_rows,
    setting_rows,
    tag_description,
)

# Per-project cache. Each imported ACD lands in cache/{project}/ with its copy
# of the ACD, the intermediate L5X, the built markup file, and the exploded tree
# (exploded/RSLogix5000Content/...). The exploded tree is the browse root.
CACHE_DIR = (Path(__file__).parent / "cache").resolve()


def project_dir(project: str) -> Path:
    """Absolute path to a project's cache folder."""
    return (CACHE_DIR / project).resolve()


def exploded_root(project: str) -> Path:
    """Browse root (the exploded tree) for a project."""
    return project_dir(project) / "exploded"


def markup_path(project: str) -> Path:
    """Path to a project's persisted content-search markup file."""
    return project_dir(project) / f"markup_{project}.xml"


def list_projects() -> list[str]:
    """Project names under cache/ that contain an exploded/ tree, sorted."""
    if not CACHE_DIR.is_dir():
        return []
    names = [
        entry.name
        for entry in CACHE_DIR.iterdir()
        if entry.is_dir() and (entry / "exploded").is_dir()
    ]
    return sorted(names, key=str.lower)


def default_project() -> str | None:
    """Most recently modified project, or None when the cache is empty."""
    projects = list_projects()
    if not projects:
        return None
    return max(projects, key=lambda p: project_dir(p).stat().st_mtime)


SETTINGS_REL = "RSLogix5000Content/Tags/Settings.xml"
RECIPE_REL = "RSLogix5000Content/Tags/Machine_Run_Recipe.xml"
MAINPROGRAM_TAGS_REL = "RSLogix5000Content/Programs/MainProgram/Tags"
CONTROLLER_TAGS_REL = "RSLogix5000Content/Tags"
PROGRAMS_REL = "RSLogix5000Content/Programs"
# Alarm tag names: faults (fault/flt) plus alarms (alarm/alm). The name filter
# is just a prefilter; non-alarm tags (no AlarmConfig block) are dropped later.
FAULT_NAME_RE = re.compile(r"(fault|flt|alarm|alm)", re.IGNORECASE)
PARAM_FIELDS = ["Parameter", "Description", "Unit", "Min", "Max"]
PREVIEW_LINE_LIMIT = 2000
MAX_SEARCH_RESULTS = 100
MAX_RECENT_SEARCHES = 10

# Content search ------------------------------------------------------------
# Each project's markup file (cache/{project}/markup_{project}.xml) concatenates
# every exploded source file, each chunk prefixed by an `@@FILE` comment that
# links the text back to the exploded file it came from. The index is built once
# per project, cached in memory, and rebuilt only when the exploded tree changes.
FILE_MARKER = "@@FILE"  # token used inside the markup comment, e.g. <!-- @@FILE rel -->
SEARCHABLE_EXTS = {".xml", ".st", ".yaml"}  # source files; derived csv/ is skipped
MAX_CONTENT_FILES = 100        # max files listed in the results dropdown
MAX_SNIPPETS_PER_FILE = 5      # max matching lines previewed per file
MAX_CONTENT_MATCHES = 2000     # global cap so a broad search can't run away
MAX_SNIPPET_LEN = 200          # truncate long matching lines for display

# Per-project cached index: {project: {"signature": <tuple>, "files": [...]}}.
_CONTENT_INDEX: dict[str, dict] = {}

CHEVRON_OPEN = "\u25BE"   # down-pointing triangle
CHEVRON_CLOSED = "\u25B8" # right-pointing triangle
ICON_DIR = "\U0001F4C1"   # folder
ICON_FILE = "\U0001F4C4"  # page


def safe_resolve(rel_path: str, root: Path) -> Path:
    """Resolve a user-supplied relative path against `root`.

    Raises ValueError if the resolved path escapes `root` (path traversal guard).
    """
    rel_path = (rel_path or "").strip().lstrip("/\\")
    candidate = (root / rel_path).resolve()
    if candidate != root and root not in candidate.parents:
        raise ValueError(f"Path escapes browse root: {rel_path!r}")
    return candidate


def list_dir(rel_path: str, root: Path):
    """Return (dirs, files) for the directory at rel_path, sorted by name.

    Each entry is a dict with 'name' and 'rel' (path relative to `root`).
    """
    directory = safe_resolve(rel_path, root)
    dirs, files = [], []
    for entry in sorted(directory.iterdir(), key=lambda p: p.name.lower()):
        rel = entry.relative_to(root).as_posix()
        item = {"name": entry.name, "rel": rel}
        if entry.is_dir():
            dirs.append(item)
        else:
            files.append(item)
    return dirs, files


def all_dirs(root: Path) -> list[str]:
    """Every directory under `root` (relative paths), for 'Expand all'."""
    return [p.relative_to(root).as_posix() for p in root.rglob("*") if p.is_dir()]


def all_files(root: Path) -> list[str]:
    """Every file under `root` (relative paths), used by the search index."""
    return sorted(p.relative_to(root).as_posix() for p in root.rglob("*") if p.is_file())


def ancestors_of(rel: str) -> list[str]:
    """Directory rel-paths that must be expanded to reveal `rel`."""
    parts = rel.split("/")
    return ["/".join(parts[: i + 1]) for i in range(len(parts) - 1)]


def build_tree_nodes(rel_path, expanded, selected, root, depth=0):
    """Recursively build indented tree rows.

    Only the children of directories present in `expanded` are rendered, so the
    tree is loaded lazily as the user expands folders.
    """
    nodes = []
    dirs, files = list_dir(rel_path, root)

    for d in dirs:
        is_open = d["rel"] in expanded
        chevron = CHEVRON_OPEN if is_open else CHEVRON_CLOSED
        nodes.append(
            html.Button(
                f"{chevron} {ICON_DIR} {d['name']}",
                id={"type": "tree-dir", "index": d["rel"]},
                n_clicks=0,
                className="tree-row tree-dir",
                style={"paddingLeft": f"{depth * 16 + 6}px"},
            )
        )
        if is_open:
            nodes.extend(build_tree_nodes(d["rel"], expanded, selected, root, depth + 1))

    for f in files:
        selected_cls = " tree-selected" if f["rel"] == selected else ""
        nodes.append(
            html.Button(
                f"{ICON_FILE} {f['name']}",
                id={"type": "tree-file", "index": f["rel"]},
                n_clicks=0,
                className=f"tree-row tree-file{selected_cls}",
                style={"paddingLeft": f"{depth * 16 + 22}px"},
            )
        )

    return nodes


def pretty_xml(path: Path) -> str:
    """Return indented XML text (falls back to raw text if it won't parse)."""
    try:
        raw = path.read_text(encoding="utf-8", errors="replace")
    except OSError as exc:
        return f"(could not read file: {exc})"

    try:
        dom = minidom.parseString(raw)
        pretty = dom.toprettyxml(indent="  ")
        text = "\n".join(line for line in pretty.splitlines() if line.strip())
    except Exception:
        text = raw

    lines = text.splitlines()
    if len(lines) > PREVIEW_LINE_LIMIT:
        lines = lines[:PREVIEW_LINE_LIMIT]
        lines.append(f"... (truncated at {PREVIEW_LINE_LIMIT} lines)")
    return "\n".join(lines)


def search_files(query: str, root: Path) -> list[str]:
    """Return file rel-paths containing `query` (case-insensitive)."""
    query = (query or "").strip()
    if not query:
        return []
    needle = query.lower()
    matcher = lambda s: needle in s.lower()
    return [f for f in all_files(root) if matcher(f)][:MAX_SEARCH_RESULTS]


# ---------------------------------------------------------------------------
# Content search index
# ---------------------------------------------------------------------------
def iter_source_files(root: Path):
    """Yield exploded source files worth searching, in stable order."""
    for path in sorted(root.rglob("*"), key=lambda p: p.as_posix().lower()):
        if path.is_file() and path.suffix.lower() in SEARCHABLE_EXTS:
            yield path


def index_signature(root: Path) -> tuple:
    """Cheap fingerprint of the exploded tree to detect when to rebuild.

    Uses (file count, newest mtime) so edits or re-explodes invalidate the
    cached index without us re-reading every file on each search.
    """
    count = 0
    newest = 0.0
    for path in iter_source_files(root):
        count += 1
        newest = max(newest, path.stat().st_mtime)
    return (count, newest)


def build_content_index(root: Path, project: str) -> list[dict]:
    """Read every source file once and persist the project's markup file.

    Returns the in-memory index (a list of {"rel", "text"}). Also writes
    `cache/{project}/markup_{project}.xml` where each file's text is preceded by
    an `<!-- @@FILE rel -->` comment, so the same content can be searched as one
    document and every hit traced back to its exploded file by its title.
    """
    files: list[dict] = []
    markup_parts: list[str] = []
    for path in iter_source_files(root):
        rel = path.relative_to(root).as_posix()
        try:
            text = path.read_text(encoding="utf-8", errors="replace")
        except OSError:
            continue
        files.append({"rel": rel, "text": text})
        markup_parts.append(f"<!-- {FILE_MARKER} {rel} -->")
        markup_parts.append(text)

    try:
        out = markup_path(project)
        out.parent.mkdir(parents=True, exist_ok=True)
        out.write_text("\n".join(markup_parts), encoding="utf-8")
    except OSError:
        pass  # search still works from the in-memory index if the write fails

    return files


def get_content_index(root: Path, project: str) -> list[dict]:
    """Return a project's cached index, rebuilding if its tree changed."""
    signature = index_signature(root)
    entry = _CONTENT_INDEX.get(project)
    if entry is None or entry["signature"] != signature:
        entry = {"signature": signature, "files": build_content_index(root, project)}
        _CONTENT_INDEX[project] = entry
    return entry["files"]


def search_content(query: str, root: Path, project: str):
    """Search file contents for `query`.

    Returns (results, matched_query) where results is a list of
    {"rel", "count", "hits": [(lineno, line), ...]} sorted by match count.
    The search runs over the cached index (built once), so it never has to
    re-read the exploded files from disk.
    """
    query = (query or "").strip()
    if not query:
        return [], None

    needle = query.lower()
    results: list[dict] = []
    total = 0
    for entry in get_content_index(root, project):
        hits: list[tuple[int, str]] = []
        count = 0
        for lineno, line in enumerate(entry["text"].splitlines(), 1):
            if needle in line.lower():
                count += 1
                total += 1
                if len(hits) < MAX_SNIPPETS_PER_FILE:
                    hits.append((lineno, line.strip()[:MAX_SNIPPET_LEN]))
                if total >= MAX_CONTENT_MATCHES:
                    break
        if count:
            results.append({"rel": entry["rel"], "count": count, "hits": hits})
        if total >= MAX_CONTENT_MATCHES:
            break

    results.sort(key=lambda r: (-r["count"], r["rel"]))
    return results[:MAX_CONTENT_FILES], query


def highlight(line: str, query: str) -> list:
    """Split `line` into text + html.Mark spans around each text match."""
    if not query:
        return [line]
    parts: list = []
    last = 0
    lower_line = line.lower()
    needle = query.lower()
    start = lower_line.find(needle)
    while start != -1:
        end = start + len(query)
        parts.append(line[last:start])
        parts.append(html.Mark(line[start:end]))
        last = end
        start = lower_line.find(needle, end)
    parts.append(line[last:])
    return parts or [line]


def real_click(ctx) -> bool:
    """True only for a genuine click.

    When the tree (or tab strip) re-renders, recreated buttons reset n_clicks to
    0/None, which re-fires pattern-matching callbacks. Those spurious triggers
    carry a falsy value and must be ignored.
    """
    return bool(ctx.triggered and ctx.triggered[0].get("value"))


def add_tab(open_tabs, rel):
    """Return a new open-tabs list with `rel` appended if not already present."""
    tabs = list(open_tabs or [])
    if rel not in tabs:
        tabs.append(rel)
    return tabs


def render_file(rel: str, root: Path):
    """Parse a file into the cache payload for preview."""
    try:
        path = safe_resolve(rel, root)
    except ValueError as exc:
        return {"xml_md": f"```\n{exc}\n```"}

    if path.suffix.lower() != ".xml":
        try:
            raw = path.read_text(encoding="utf-8", errors="replace")
        except OSError as exc:
            raw = f"(could not read file: {exc})"
        return {"xml_md": f"```\n{raw}\n```"}

    return {"xml_md": f"```xml\n{pretty_xml(path)}\n```"}


THEME = {
    "primaryColor": "blue",
    "fontFamily": "system-ui, -apple-system, 'Segoe UI', sans-serif",
    "fontFamilyMonospace": "'Cascadia Code', Consolas, 'Courier New', monospace",
}

NAVBAR_WIDTH = 320


# Background callback manager: ACD conversion takes minutes, so it runs off the
# request thread to keep the server responsive while importing.
_background_manager = DiskcacheManager(diskcache.Cache(str(CACHE_DIR / ".dash_cache")))

app = dash.Dash(
    __name__,
    suppress_callback_exceptions=True,
    background_callback_manager=_background_manager,
)
app.title = "Studio5000 XML Viewer"


# ---------------------------------------------------------------------------
# Layout pieces
# ---------------------------------------------------------------------------
def header():
    return dmc.AppShellHeader(
        dmc.Group(
            [
                dmc.Group(
                    [
                        html.Img(
                            src=dash.get_asset_url("trunorth-logo.png"),
                            style={"height": "44px"},
                        ),
                        dmc.Title("Studio5000 XML Viewer", order=4),
                    ],
                    gap="sm",
                    align="center",
                ),
            ],
            justify="space-between",
            align="center",
            h="100%",
            px="md",
        ),
    )


def navbar():
    return dmc.AppShellNavbar(
        [
            html.Div(
                [
                    html.Div(
                        [
                            html.Div("Import .ACD", className="sidebar-import-label"),
                            dcc.Upload(
                                id="acd-upload",
                                accept=".acd,.ACD",
                                max_size=-1,
                                children=html.Div(
                                    [
                                        DashIconify(icon="tabler:upload", width=18),
                                        html.Span("Drop an .ACD here or click to browse"),
                                    ],
                                    className="acd-dropzone-inner",
                                ),
                                className="acd-dropzone",
                            ),
                            dmc.Select(
                                id="project-switcher",
                                placeholder="No project",
                                data=[],
                                value=None,
                                leftSection=DashIconify(icon="tabler:folder"),
                                comboboxProps={"withinPortal": False},
                                size="sm",
                                allowDeselect=False,
                                mt=8,
                            ),
                            dcc.Loading(
                                html.Div(id="import-status", className="import-status"),
                                type="dot",
                            ),
                        ],
                        className="sidebar-import",
                    ),
                    html.Div(
                        "Project Files",
                        className="sidebar-title",
                        style={"padding": "14px 12px 12px 12px", "marginBottom": "12px"},
                    ),
                    html.Div(
                        [
                            dmc.Autocomplete(
                                id="search",
                                placeholder="Search files\u2026",
                                leftSection=DashIconify(icon="tabler:search"),
                                rightSection=dmc.ActionIcon(
                                    DashIconify(icon="tabler:x", width=14),
                                    id="clear-search",
                                    n_clicks=0,
                                    variant="subtle",
                                    size="sm",
                                    color="gray",
                                ),
                                data=[],
                                comboboxProps={"withinPortal": False},
                                size="sm",
                            ),
                            html.Div(id="search-results", style={"display": "none"}),
                        ],
                        className="sidebar-search-block",
                        style={"position": "relative", "padding": "0 8px 6px 8px"},
                    ),
                    html.Div(
                        [
                            dmc.TextInput(
                                id="content-search",
                                placeholder="Search file contents\u2026",
                                leftSection=DashIconify(icon="tabler:file-search"),
                                rightSection=dmc.ActionIcon(
                                    DashIconify(icon="tabler:x", width=14),
                                    id="clear-content-search",
                                    n_clicks=0,
                                    variant="subtle",
                                    size="sm",
                                    color="gray",
                                ),
                                size="sm",
                            ),
                            html.Div(
                                dmc.Button(
                                    "Open all matches",
                                    id="open-all-btn",
                                    n_clicks=0,
                                    leftSection=DashIconify(icon="tabler:folders"),
                                    variant="light",
                                    size="xs",
                                    fullWidth=True,
                                ),
                                id="open-all-wrap",
                                style={"display": "none"},
                            ),
                            html.Div(
                                id="content-search-results",
                                style={"display": "none"},
                            ),
                        ],
                        style={"position": "relative", "padding": "0 8px 6px 8px"},
                    ),
                    dmc.Group(
                        [
                            dmc.Button(
                                "Collapse Tree",
                                id="collapse-all",
                                n_clicks=0,
                                leftSection=DashIconify(
                                    icon="tabler:layout-sidebar-left-collapse"
                                ),
                                variant="default",
                                size="xs",
                                fullWidth=True,
                            ),
                            dmc.Button(
                                "Expand Tree",
                                id="expand-all",
                                n_clicks=0,
                                leftSection=DashIconify(
                                    icon="tabler:layout-sidebar-right-collapse"
                                ),
                                variant="default",
                                size="xs",
                                fullWidth=True,
                            ),
                        ],
                        grow=True,
                        gap="xs",
                        px="sm",
                        pb="xs",
                    ),
                    dmc.ScrollArea(
                        html.Div(id="tree", className="tree"),
                        style={"flex": "1 1 0", "minHeight": "0"},
                        type="auto",
                    ),
                    html.Div(
                        dmc.Button(
                            "SD/DV Format",
                            id="export-sd-tables-btn",
                            n_clicks=0,
                            leftSection=DashIconify(icon="tabler:download"),
                            variant="filled",
                            color="blue",
                            size="sm",
                            fullWidth=True,
                        ),
                        className="sidebar-export",
                    ),
                ],
                id="nav-body",
                className="nav-body",
            ),
        ],
        style={"display": "flex", "flexDirection": "column", "minHeight": "0"},
        p=0,
    )


def xml_block():
    return html.Div(
        [
            html.Div(
                [
                    html.Span("XML", className="section-label"),
                    html.Span(id="xml-path", className="file-path"),
                ],
                className="section-title",
            ),
            dcc.Markdown(id="xml-preview", className="xml-view"),
        ],
        id="xml-block",
        className="preview-block",
    )


def empty_state():
    return html.Div(
        dmc.Stack(
            [
                DashIconify(icon="tabler:file-text", width=56, color="#c0c7cf"),
                dmc.Text("No file open", fw=600, c="dimmed"),
                dmc.Text(
                    "Pick a file from the tree, or use search to open one.",
                    size="sm",
                    c="dimmed",
                ),
            ],
            align="center",
            gap=6,
        ),
        id="empty-state",
        className="empty-state",
    )


def main_panel():
    return dmc.AppShellMain(
        [
            dmc.Tabs(
                id="file-tabs",
                value=None,
                children=[],
                variant="outline",
                className="file-tabs",
            ),
            html.Div(
                xml_block(),
                id="preview-container",
            ),
            empty_state(),
        ],
        className="app-main",
    )


def fault_range_row(label: str, key: str, min_default: int, max_default: int):
    """A labelled Min/Max severity range row for the Fault Configuration modal."""
    return html.Div(
        [
            dmc.Text(label, fw=600, size="sm"),
            dmc.Group(
                [
                    dmc.NumberInput(
                        id=f"fault-{key}-min",
                        label="Min",
                        value=min_default,
                        allowDecimal=False,
                        className="fault-range-input",
                        style={"flex": "1 1 0"},
                    ),
                    dmc.NumberInput(
                        id=f"fault-{key}-max",
                        label="Max",
                        value=max_default,
                        allowDecimal=False,
                        className="fault-range-input",
                        style={"flex": "1 1 0"},
                    ),
                ],
                grow=True,
                gap="xs",
            ),
        ],
    )


app.layout = dmc.MantineProvider(
    forceColorScheme="light",
    theme=THEME,
    children=dmc.AppShell(
        [
            dcc.Store(id="active-project", storage_type="local", data=default_project()),
            dcc.Store(id="expanded", data=["RSLogix5000Content"]),
            dcc.Store(id="open-tabs", storage_type="local", data=[]),
            dcc.Store(id="active-tab", storage_type="local", data=None),
            dcc.Store(id="recent-searches", storage_type="local", data=[]),
            dcc.Store(id="content-search-rels", data=[]),
            dcc.Store(id="file-cache", storage_type="memory", data={}),
            dcc.Download(id="download-sd-tables"),
            dmc.Modal(
                id="fault-config-modal",
                title=html.Span("Alarm Configuration", className="fault-config-title"),
                centered=True,
                children=[
                    dmc.Stack(
                        [
                            fault_range_row("Faults", "faults", 500, 99999),
                            fault_range_row("Warnings", "warnings", 200, 399),
                            fault_range_row("Notifications", "notifications", 0, 199),
                            html.Div(
                                [
                                    dmc.Button(
                                        "Generate Excel",
                                        id="fault-config-generate",
                                        n_clicks=0,
                                        color="blue",
                                        style={"flex": "1 1 0"},
                                    ),
                                    dmc.Button(
                                        "Cancel",
                                        id="fault-config-cancel",
                                        n_clicks=0,
                                        variant="default",
                                        style={"flex": "1 1 0"},
                                    ),
                                ],
                                style={"display": "flex", "gap": "8px"},
                            ),
                        ],
                        gap="sm",
                    ),
                ],
                opened=False,
            ),
            dmc.Modal(
                id="open-all-modal",
                title="Open all matching files?",
                centered=True,
                children=[
                    dmc.Text(id="open-all-text"),
                    dmc.Group(
                        [
                            dmc.Button("Cancel", id="open-all-cancel", variant="default"),
                            dmc.Button("Open all", id="open-all-confirm", color="blue"),
                        ],
                        justify="flex-end",
                        mt="md",
                    ),
                ],
                opened=False,
            ),
            header(),
            navbar(),
            main_panel(),
        ],
        id="app-shell",
        header={"height": 56},
        navbar={
            "width": NAVBAR_WIDTH,
            "breakpoint": "sm",
            "collapsed": {"desktop": False, "mobile": False},
        },
        padding="md",
    ),
)


# ---------------------------------------------------------------------------
# Projects (import + switcher)
# ---------------------------------------------------------------------------
TREE_RESET = ["RSLogix5000Content"]


@app.callback(
    Output("project-switcher", "data"),
    Output("project-switcher", "value"),
    Input("active-project", "data"),
)
def sync_project_switcher(active):
    """Keep the switcher's options and selection in step with the active project."""
    data = [{"value": p, "label": p} for p in list_projects()]
    return data, active


@app.callback(
    Output("active-project", "data", allow_duplicate=True),
    Output("expanded", "data", allow_duplicate=True),
    Output("open-tabs", "data", allow_duplicate=True),
    Output("active-tab", "data", allow_duplicate=True),
    Output("file-cache", "data", allow_duplicate=True),
    Input("project-switcher", "value"),
    State("active-project", "data"),
    prevent_initial_call=True,
)
def switch_project(value, active):
    """Switch projects: tabs/cache/expanded refer to the old project, so reset."""
    if not value or value == active:
        return (dash.no_update,) * 5
    return value, TREE_RESET, [], None, {}


@app.callback(
    Output("active-project", "data", allow_duplicate=True),
    Output("import-status", "children"),
    Output("expanded", "data", allow_duplicate=True),
    Output("open-tabs", "data", allow_duplicate=True),
    Output("active-tab", "data", allow_duplicate=True),
    Output("file-cache", "data", allow_duplicate=True),
    Input("acd-upload", "contents"),
    State("acd-upload", "filename"),
    background=True,
    running=[
        (Output("acd-upload", "disabled"), True, False),
        (Output("project-switcher", "disabled"), True, False),
    ],
    prevent_initial_call=True,
)
def handle_import(contents, filename):
    """Decode the upload, run the ACD pipeline, then activate the new project."""
    if not contents:
        return (dash.no_update,) * 6

    no_change = (dash.no_update, dash.no_update, dash.no_update, dash.no_update)
    try:
        _header, b64 = contents.split(",", 1)
        acd_bytes = base64.b64decode(b64)
    except (ValueError, base64.binascii.Error):
        return (dash.no_update, "Could not read the uploaded file.", *no_change)

    try:
        project = import_acd(acd_bytes, filename)
        get_content_index(exploded_root(project), project)  # prebuild markup
    except PipelineError as exc:
        return (dash.no_update, str(exc), *no_change)

    status = f"Imported {project}."
    return project, status, TREE_RESET, [], None, {}


# ---------------------------------------------------------------------------
# Search
# ---------------------------------------------------------------------------
@app.callback(
    Output("search", "data"),
    Input("recent-searches", "data"),
)
def sync_search_history(recent):
    return recent or []


@app.callback(
    Output("search", "value", allow_duplicate=True),
    Input("clear-search", "n_clicks"),
    prevent_initial_call=True,
)
def clear_search(_clicks):
    if not real_click(dash.callback_context):
        return dash.no_update
    return ""


@app.callback(
    Output("search-results", "children"),
    Output("search-results", "style"),
    Input("search", "value"),
    State("active-project", "data"),
)
def do_search(query, project):
    hidden = {"display": "none"}
    if not (query or "").strip() or not project:
        return [], hidden

    matches = search_files(query, exploded_root(project))
    shown = {
        "position": "absolute",
        "left": "8px",
        "right": "8px",
        "top": "100%",
        "zIndex": 50,
        "background": "#fff",
        "border": "1px solid #d0d7de",
        "borderRadius": "6px",
        "maxHeight": "55vh",
        "overflowY": "auto",
        "boxShadow": "0 6px 18px rgba(0,0,0,.12)",
    }
    if not matches:
        return [html.Div("No matches", className="search-empty")], shown

    results = [
        html.Button(
            [
                html.Div(rel.rsplit("/", 1)[-1], className="sr-name"),
                html.Div(rel, className="sr-path"),
            ],
            id={"type": "search-result", "index": rel},
            n_clicks=0,
            className="search-result",
        )
        for rel in matches
    ]
    return results, shown


@app.callback(
    Output("open-tabs", "data", allow_duplicate=True),
    Output("active-tab", "data", allow_duplicate=True),
    Output("recent-searches", "data", allow_duplicate=True),
    Output("search", "value"),
    Input({"type": "search-result", "index": dash.ALL}, "n_clicks"),
    State("open-tabs", "data"),
    State("recent-searches", "data"),
    State("search", "value"),
    prevent_initial_call=True,
)
def pick_search_result(_clicks, open_tabs, recent, query):
    ctx = dash.callback_context
    tid = ctx.triggered_id
    if not isinstance(tid, dict) or tid.get("type") != "search-result" or not real_click(ctx):
        return dash.no_update, dash.no_update, dash.no_update, dash.no_update
    rel = tid["index"]
    query = (query or "").strip()
    recent = [q for q in (recent or []) if q != query]
    if query:
        recent.insert(0, query)
    recent = recent[:MAX_RECENT_SEARCHES]
    return add_tab(open_tabs, rel), rel, recent, ""


# ---------------------------------------------------------------------------
# Content search (search inside files)
# ---------------------------------------------------------------------------
DROPDOWN_STYLE = {
    "position": "absolute",
    "left": "8px",
    "right": "auto",
    "top": "100%",
    "width": "40vw",
    "minWidth": "360px",
    "zIndex": 50,
    "background": "#fff",
    "border": "1px solid #d0d7de",
    "borderRadius": "6px",
    "maxHeight": "55vh",
    "overflowY": "auto",
    "boxShadow": "0 6px 18px rgba(0,0,0,.12)",
}


@app.callback(
    Output("content-search", "value", allow_duplicate=True),
    Input("clear-content-search", "n_clicks"),
    prevent_initial_call=True,
)
def clear_content_search(_clicks):
    if not real_click(dash.callback_context):
        return dash.no_update
    return ""


@app.callback(
    Output("content-search-results", "children"),
    Output("content-search-results", "style"),
    Output("open-all-wrap", "style"),
    Output("open-all-btn", "children"),
    Output("content-search-rels", "data"),
    Input("content-search", "value"),
    State("active-project", "data"),
)
def do_content_search(query, project):
    hidden = {"display": "none"}
    btn_hidden = {"display": "none"}
    if not (query or "").strip() or not project:
        return [], hidden, btn_hidden, dash.no_update, []

    results, matched_query = search_content(query, exploded_root(project), project)
    if not results:
        return [html.Div("No matches", className="search-empty")], DROPDOWN_STYLE, btn_hidden, dash.no_update, []

    rels = [r["rel"] for r in results]
    btn_shown = {"padding": "0 8px 6px 8px"}
    btn_label = f"Open all {len(rels)} file{'s' if len(rels) != 1 else ''}"

    total = sum(r["count"] for r in results)
    children = [
        html.Div(
            f"{total} match{'es' if total != 1 else ''} in {len(results)} file"
            f"{'s' if len(results) != 1 else ''}",
            className="cs-summary",
        )
    ]
    for r in results:
        rel = r["rel"]
        snippets = [
            html.Div(
                [
                    html.Span(f"{lineno}", className="cs-lineno"),
                    html.Span(highlight(line, matched_query), className="cs-line"),
                ],
                className="cs-snippet",
            )
            for lineno, line in r["hits"]
        ]
        children.append(
            html.Button(
                [
                    html.Div(
                        [
                            html.Span(rel.rsplit("/", 1)[-1], className="sr-name"),
                            html.Span(str(r["count"]), className="cs-count"),
                        ],
                        className="cs-head",
                    ),
                    html.Div(rel, className="sr-path"),
                    html.Div(snippets, className="cs-snippets"),
                ],
                id={"type": "content-result", "index": rel},
                n_clicks=0,
                className="content-result",
            )
        )
    return children, DROPDOWN_STYLE, btn_shown, btn_label, rels


@app.callback(
    Output("open-tabs", "data", allow_duplicate=True),
    Output("active-tab", "data", allow_duplicate=True),
    Output("content-search", "value"),
    Input({"type": "content-result", "index": dash.ALL}, "n_clicks"),
    State("open-tabs", "data"),
    prevent_initial_call=True,
)
def pick_content_result(_clicks, open_tabs):
    ctx = dash.callback_context
    tid = ctx.triggered_id
    if not isinstance(tid, dict) or tid.get("type") != "content-result" or not real_click(ctx):
        return dash.no_update, dash.no_update, dash.no_update
    rel = tid["index"]
    return add_tab(open_tabs, rel), rel, ""


@app.callback(
    Output("open-all-modal", "opened"),
    Output("open-all-text", "children"),
    Input("open-all-btn", "n_clicks"),
    Input("open-all-confirm", "n_clicks"),
    Input("open-all-cancel", "n_clicks"),
    State("content-search-rels", "data"),
    prevent_initial_call=True,
)
def toggle_open_all_modal(_open, _confirm, _cancel, rels):
    ctx = dash.callback_context
    tid = ctx.triggered_id
    if tid == "open-all-btn" and real_click(ctx):
        count = len(rels or [])
        return True, (
            f"This will open {count} file{'s' if count != 1 else ''} in new tabs."
        )
    if tid in ("open-all-confirm", "open-all-cancel"):
        return False, dash.no_update
    return dash.no_update, dash.no_update


@app.callback(
    Output("open-tabs", "data", allow_duplicate=True),
    Output("active-tab", "data", allow_duplicate=True),
    Output("content-search", "value", allow_duplicate=True),
    Input("open-all-confirm", "n_clicks"),
    State("content-search-rels", "data"),
    State("open-tabs", "data"),
    prevent_initial_call=True,
)
def open_all_confirmed(_clicks, rels, open_tabs):
    ctx = dash.callback_context
    if not real_click(ctx):
        return dash.no_update, dash.no_update, dash.no_update
    rels = rels or []
    if not rels:
        return dash.no_update, dash.no_update, dash.no_update
    tabs = list(open_tabs or [])
    for rel in rels:
        if rel not in tabs:
            tabs.append(rel)
    return tabs, rels[0], ""


# ---------------------------------------------------------------------------
# Tree
# ---------------------------------------------------------------------------
@app.callback(
    Output("expanded", "data"),
    Input({"type": "tree-dir", "index": dash.ALL}, "n_clicks"),
    Input("expand-all", "n_clicks"),
    Input("collapse-all", "n_clicks"),
    State("expanded", "data"),
    State("active-project", "data"),
    prevent_initial_call=True,
)
def update_expanded(_dir_clicks, _expand, _collapse, expanded, project):
    ctx = dash.callback_context
    tid = ctx.triggered_id
    if tid == "expand-all":
        return all_dirs(exploded_root(project)) if project else dash.no_update
    if tid == "collapse-all":
        return []
    # A folder row toggle: ignore spurious re-render triggers (n_clicks reset).
    if not isinstance(tid, dict) or tid.get("type") != "tree-dir" or not real_click(ctx):
        return dash.no_update
    expanded = list(expanded or [])
    rel = tid["index"]
    if rel in expanded:
        expanded.remove(rel)
    else:
        expanded.append(rel)
    return expanded


@app.callback(
    Output("open-tabs", "data", allow_duplicate=True),
    Output("active-tab", "data", allow_duplicate=True),
    Input({"type": "tree-file", "index": dash.ALL}, "n_clicks"),
    State("open-tabs", "data"),
    prevent_initial_call=True,
)
def open_from_tree(_clicks, open_tabs):
    ctx = dash.callback_context
    tid = ctx.triggered_id
    if not isinstance(tid, dict) or tid.get("type") != "tree-file" or not real_click(ctx):
        return dash.no_update, dash.no_update
    rel = tid["index"]
    return add_tab(open_tabs, rel), rel


@app.callback(
    Output("tree", "children"),
    Input("expanded", "data"),
    Input("active-tab", "data"),
    Input("active-project", "data"),
)
def render_tree(expanded, active, project):
    if not project:
        return html.Div(
            "No project loaded. Import an .ACD to get started.",
            className="tree-empty",
        )
    try:
        return build_tree_nodes("", set(expanded or []), active, exploded_root(project))
    except (ValueError, OSError) as exc:
        return html.Div(f"Error: {exc}", className="status-error")


# ---------------------------------------------------------------------------
# Tabs
# ---------------------------------------------------------------------------
@app.callback(
    Output("file-tabs", "children"),
    Output("file-tabs", "value"),
    Input("open-tabs", "data"),
    Input("active-tab", "data"),
)
def render_tabs(open_tabs, active):
    open_tabs = open_tabs or []
    tabs = [
        dmc.TabsTab(
            rel.rsplit("/", 1)[-1],
            value=rel,
            rightSection=dmc.ActionIcon(
                DashIconify(icon="tabler:x", width=12),
                id={"type": "tab-close", "index": rel},
                n_clicks=0,
                variant="subtle",
                color="gray",
                size="xs",
            ),
        )
        for rel in open_tabs
    ]
    open_file_items = [
        dmc.MenuItem(
            [
                html.Div(rel.rsplit("/", 1)[-1], className="of-name"),
                html.Div(rel, className="of-path"),
            ],
            id={"type": "open-file-menu-item", "index": rel},
            n_clicks=0,
        )
        for rel in open_tabs
    ] or [dmc.MenuItem("No open files", disabled=True)]
    controls = dmc.Group(
        [
            dmc.Menu(
                [
                    dmc.MenuTarget(
                        dmc.ActionIcon(
                            DashIconify(icon="tabler:chevron-down"),
                            id="open-files-menu-btn",
                            variant="default",
                            size="sm",
                            disabled=not open_tabs,
                        )
                    ),
                    dmc.MenuDropdown(open_file_items, className="open-files-menu"),
                ],
                position="bottom-end",
                withinPortal=True,
            ),
            dmc.Tooltip(
                dmc.ActionIcon(
                    DashIconify(icon="tabler:trash"),
                    id="clear-open-files",
                    n_clicks=0,
                    variant="default",
                    color="red",
                    size="sm",
                    disabled=not open_tabs,
                ),
                label="Clear open files",
            ),
        ],
        gap=4,
        className="file-tab-controls",
    )
    return [html.Div([dmc.TabsList(tabs), controls], className="file-tabs-row")], active


@app.callback(
    Output("active-tab", "data", allow_duplicate=True),
    Input("file-tabs", "value"),
    State("active-tab", "data"),
    prevent_initial_call=True,
)
def switch_tab(value, active):
    if value == active:
        return dash.no_update
    return value


@app.callback(
    Output("active-tab", "data", allow_duplicate=True),
    Input({"type": "open-file-menu-item", "index": dash.ALL}, "n_clicks"),
    prevent_initial_call=True,
)
def open_from_open_files_menu(_clicks):
    ctx = dash.callback_context
    tid = ctx.triggered_id
    if not isinstance(tid, dict) or tid.get("type") != "open-file-menu-item" or not real_click(ctx):
        return dash.no_update
    return tid["index"]


@app.callback(
    Output("open-tabs", "data", allow_duplicate=True),
    Output("active-tab", "data", allow_duplicate=True),
    Input("clear-open-files", "n_clicks"),
    prevent_initial_call=True,
)
def clear_open_files(_clicks):
    if not real_click(dash.callback_context):
        return dash.no_update, dash.no_update
    return [], None


@app.callback(
    Output("open-tabs", "data", allow_duplicate=True),
    Output("active-tab", "data", allow_duplicate=True),
    Input({"type": "tab-close", "index": dash.ALL}, "n_clicks"),
    State("open-tabs", "data"),
    State("active-tab", "data"),
    prevent_initial_call=True,
)
def close_tab(_clicks, open_tabs, active):
    ctx = dash.callback_context
    tid = ctx.triggered_id
    if not isinstance(tid, dict) or tid.get("type") != "tab-close" or not real_click(ctx):
        return dash.no_update, dash.no_update
    rel = tid["index"]
    tabs = list(open_tabs or [])
    if rel not in tabs:
        return dash.no_update, dash.no_update
    idx = tabs.index(rel)
    tabs.remove(rel)
    if active != rel:
        return tabs, dash.no_update
    if not tabs:
        return tabs, None
    neighbor = tabs[idx - 1] if idx > 0 else tabs[0]
    return tabs, neighbor


# ---------------------------------------------------------------------------
# Active preview (cached)
# ---------------------------------------------------------------------------
@app.callback(
    Output("xml-preview", "children"),
    Output("file-cache", "data"),
    Output("xml-path", "children"),
    Input("active-tab", "data"),
    State("file-cache", "data"),
    State("active-project", "data"),
)
def render_active(rel, cache, project):
    if not rel or not project:
        return "", dash.no_update, ""

    cache = cache or {}
    hit = cache.get(rel)
    if hit is None:
        hit = render_file(rel, exploded_root(project))
        cache = {**cache, rel: hit}
        cache_out = cache
    else:
        cache_out = dash.no_update

    return hit["xml_md"], cache_out, rel


@app.callback(
    Output("preview-container", "style"),
    Output("empty-state", "style"),
    Input("active-tab", "data"),
)
def toggle_empty(rel):
    """Show the preview panes when a file is open, the empty state otherwise."""
    if rel:
        return {"display": "flex"}, {"display": "none"}
    return {"display": "none"}, {"display": "flex"}


def _params_dataframe(rel: str, row_func, root: Path, columns=PARAM_FIELDS) -> pd.DataFrame | None:
    """Parse a tag file and extract rows into a table with the given columns."""
    try:
        path = safe_resolve(rel, root)
    except ValueError:
        return None
    if not path.is_file():
        return None

    tree = ET.parse(path)
    rows = list(row_func(tree.getroot()))
    return pd.DataFrame(rows, columns=columns)


def _make_desc_resolver(search_rels: list[str], root: Path):
    """Return f(tag_ref)->description, caching lookups across the given dirs.

    A tag reference like "actCanInsertVac.Desc" is reduced to its tag name
    ("actCanInsertVac") and resolved to that tag file's <Description> CDATA.
    """
    search_dirs = []
    for rel in search_rels:
        try:
            directory = safe_resolve(rel, root)
        except ValueError:
            continue
        if directory.is_dir():
            search_dirs.append(directory)

    cache: dict[str, str] = {}

    def resolve(ref: str) -> str:
        name = (ref or "").split(".")[0].strip()
        if not name:
            return ""
        if name not in cache:
            cache[name] = tag_description(name, search_dirs)
        return cache[name]

    return resolve


FAULT_CATEGORIES = ("Faults", "Warnings", "Notifications")


def _coerce_code(value, default: int) -> int:
    """Coerce a modal NumberInput value to an int severity code."""
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _classify_severity(severity, ranges: dict[str, tuple[int, int]]) -> str:
    """Map a fault severity to the category whose [min, max] range contains it.

    `ranges` maps each category name to an inclusive (min, max) severity range
    (from the Alarm Configuration modal). Faults whose severity matches no range
    (or have no severity at all) default to "Faults".
    """
    if severity is not None:
        for category in ("Notifications", "Warnings", "Faults"):
            low, high = ranges[category]
            if low <= severity <= high:
                return category
    return "Faults"


def _alarm_scopes(root: Path) -> list[tuple[str, str]]:
    """Return (scope_label, tags_rel) for the controller and every program scope.

    Controller tags map to the "Global" scope; each program's Tags folder maps
    to a scope labelled with the program name (e.g. "MainProgram"). Scopes whose
    Tags folder is missing are skipped.
    """
    scopes: list[tuple[str, str]] = [("Global", CONTROLLER_TAGS_REL)]
    try:
        programs_dir = safe_resolve(PROGRAMS_REL, root)
    except ValueError:
        return scopes
    if programs_dir.is_dir():
        for program in sorted(programs_dir.iterdir(), key=lambda p: p.name.lower()):
            if (program / "Tags").is_dir():
                scopes.append((program.name, f"{PROGRAMS_REL}/{program.name}/Tags"))
    return scopes


def _faults_dataframes(ranges: dict[str, tuple[int, int]], root: Path) -> dict[str, pd.DataFrame] | None:
    """Collate fault alarm tags grouped by severity classification.

    Scans the controller scope and every program scope. Within each scope, lists
    files whose name contains "fault"/"flt"/"alarm"/"alm", parses each, keeps only
    true alarm tags (those with an AlarmConfig block), tags them with their scope,
    and groups them into the "Faults", "Warnings", and "Notifications" buckets
    based on each tag's severity range. Associated %TagN placeholders resolve
    against the tag's own scope, then MainProgram, then Global. Always returns a
    frame for every category (possibly empty); None only when no scope exists.
    """
    scopes = _alarm_scopes(root)
    grouped: dict[str, list[dict]] = {category: [] for category in FAULT_CATEGORIES}
    found_any = False

    for scope_label, tags_rel in scopes:
        try:
            tags_dir = safe_resolve(tags_rel, root)
        except ValueError:
            continue
        if not tags_dir.is_dir():
            continue
        found_any = True

        resolve_desc = _make_desc_resolver(
            [tags_rel, MAINPROGRAM_TAGS_REL, CONTROLLER_TAGS_REL], root
        )
        candidates = sorted(
            (p for p in tags_dir.glob("*.xml") if FAULT_NAME_RE.search(p.stem)),
            key=lambda p: p.name.lower(),
        )
        for path in candidates:
            try:
                tree = ET.parse(path)
            except ET.ParseError:
                continue
            row = fault_row(tree.getroot(), resolve_desc=resolve_desc, scope=scope_label)
            if row is None:
                continue
            category = _classify_severity(row.get("Severity"), ranges)
            grouped[category].append(row)

    if not found_any:
        return None

    return {
        category: pd.DataFrame(rows, columns=FAULT_FIELDS)
        for category, rows in grouped.items()
    }


def _style_sd_sheet(worksheet) -> None:
    """Apply shared SD Tables workbook styling to one worksheet."""
    header_fill = PatternFill(fill_type="solid", fgColor="D9D9D9")
    index_fill = PatternFill(fill_type="solid", fgColor="F2F2F2")
    black_side = Side(style="thin", color="000000")
    header_border = Border(
        left=black_side,
        right=black_side,
        top=black_side,
        bottom=black_side,
    )
    index_border = Border(right=black_side)
    middle_alignment = Alignment(horizontal="center", vertical="center")

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.border = header_border
        cell.font = Font(bold=True)
        cell.alignment = middle_alignment

    for cell in worksheet["A"]:
        cell.fill = index_fill
        cell.border = index_border
        cell.alignment = middle_alignment

    for column_cells in worksheet.columns:
        max_len = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        column_letter = get_column_letter(column_cells[0].column)
        worksheet.column_dimensions[column_letter].width = max_len + 2


def _allow_sd_workbook_edits(workbook) -> None:
    """Ensure generated workbook opens editable by default."""
    workbook.security.lockStructure = False
    workbook.security.lockWindows = False
    workbook.security.lockRevision = False
    for worksheet in workbook.worksheets:
        worksheet.protection.disable()


def _add_index_column(df: pd.DataFrame) -> pd.DataFrame:
    """Return a copy with a 1-based Index column first."""
    out = df.copy()
    out.insert(0, "Index", range(1, len(out) + 1))
    return out


@app.callback(
    Output("fault-config-modal", "opened"),
    Input("export-sd-tables-btn", "n_clicks"),
    Input("fault-config-cancel", "n_clicks"),
    Input("fault-config-generate", "n_clicks"),
    prevent_initial_call=True,
)
def toggle_fault_config_modal(_open, _cancel, _generate):
    """Open the Alarm Configuration modal from the toolbar, close on action."""
    ctx = dash.callback_context
    tid = ctx.triggered_id
    if tid == "export-sd-tables-btn" and real_click(ctx):
        return True
    if tid in ("fault-config-cancel", "fault-config-generate"):
        return False
    return dash.no_update


@app.callback(
    Output("download-sd-tables", "data"),
    Input("fault-config-generate", "n_clicks"),
    State("fault-notifications-min", "value"),
    State("fault-notifications-max", "value"),
    State("fault-warnings-min", "value"),
    State("fault-warnings-max", "value"),
    State("fault-faults-min", "value"),
    State("fault-faults-max", "value"),
    State("active-project", "data"),
    prevent_initial_call=True,
)
def export_sd_tables(
    _clicks,
    notif_min,
    notif_max,
    warn_min,
    warn_max,
    faults_min,
    faults_max,
    project,
):
    """Build one SD Tables workbook with each table on its own sheet.

    Settings and Recipes are exported unchanged. Faults are split into
    the "Faults", "Warnings" and "Notifications" sheets according to the severity
    ranges entered in the Fault Configuration modal.
    """
    if not real_click(dash.callback_context) or not project:
        return dash.no_update

    root = exploded_root(project)
    ranges = {
        "Notifications": (_coerce_code(notif_min, 0), _coerce_code(notif_max, 199)),
        "Warnings": (_coerce_code(warn_min, 200), _coerce_code(warn_max, 399)),
        "Faults": (_coerce_code(faults_min, 500), _coerce_code(faults_max, 99999)),
    }

    settings = _params_dataframe(SETTINGS_REL, setting_rows, root)
    recipes = _params_dataframe(RECIPE_REL, recipe_rows, root)
    fault_frames = _faults_dataframes(ranges, root)

    if any(x is None for x in (settings, recipes, fault_frames)):
        return dash.no_update

    tables = {
        "Settings": settings,
        "Recipes": recipes,
        "Faults": fault_frames["Faults"],
        "Warnings": fault_frames["Warnings"],
        "Notifications": fault_frames["Notifications"],
    }

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        for sheet_name, df in tables.items():
            _add_index_column(df).to_excel(writer, sheet_name=sheet_name, index=False)
            _style_sd_sheet(writer.sheets[sheet_name])
        _allow_sd_workbook_edits(writer.book)

    return dcc.send_bytes(buffer.getvalue(), "sd_tables.xlsx")


if __name__ == "__main__":
    app.run(debug=True)
